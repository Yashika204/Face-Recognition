from flask import Flask, render_template, request, redirect, url_for,send_file
from flask_sqlalchemy import SQLAlchemy
import face_recognition
import cv2
import openpyxl
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "Name"
ws['B1'] = "Date"
ws['C1'] = "Present"
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:@localhost/faces'
db = SQLAlchemy(app)

class Face(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    s_id = db.Column(db.String(30),nullable = False)
    division = db.Column(db.String(30),nullable = False)
    department = db.Column(db.String(30),nullable = False)
    year = db.Column(db.String(30),nullable = False)
    encoding = db.Column(db.PickleType, nullable=False)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add', methods=['GET', 'POST'])
def add_face():
    if request.method == 'POST':
        name = request.form['name']
        s_id = request.form['student_id']
        div = request.form['division']
        department = request.form['department']
        year = request.form['year']
        video = cv2.VideoCapture(0)
        encoding = None
        while True:
            ret, frame = video.read()
            if not ret:
                break
            small_frame = cv2.resize(frame, (0, 0),None, 0.25,0.25)
            black_frame = cv2.cvtColor(small_frame,cv2.COLOR_BGR2RGB)
            rgb_small_frame = black_frame[:, :, ::-1]
            face_locations = face_recognition.face_locations(rgb_small_frame)
            if len(face_locations) > 0:
                face_Encodings = face_recognition.face_encodings(rgb_small_frame,face_locations)[0]
                encoding = face_Encodings[0]
                break
            cv2.imshow('Video', frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        video.release()
        cv2.destroyAllWindows()
        if encoding is not None:
            new_face = Face(name=name,s_id = s_id, division = div , department=department, year = year, encoding=encoding)
            db.session.add(new_face)
            db.session.commit()
            #return redirect(url_for('index'))
            return ("student added")    
        else:
            return("Face Not Found")
    
    return render_template('add.html')

@app.route('/take-attendance')
def take_attendance():
    knownencodings = []
    encodings = []
    students = Face.query.all()
    present_students = []
    today = datetime.date.today()
    video = cv2.VideoCapture(0)
    while True:
        ret, frame = video.read()
        if not ret:
            break
        small_frame = cv2.resize(frame, (0, 0),None, 0.25,0.25)
        black_frame = cv2.cvtColor(small_frame,cv2.COLOR_BGR2RGB)
        rgb_small_frame = black_frame[:, :, ::-1]
        face_locations = face_recognition.face_locations(rgb_small_frame)
        if len(face_locations) > 0:
            face_encodings = face_recognition.face_encodings(rgb_small_frame,face_locations)
            for student in students:
                #knownencodings.append(student.encoding)
                for face_encoding in face_encodings:
                    #matches = Face.query.filter(Face.encoding == face_encoding).all()
                    #matches = face_recognition.compare_faces([encodeListKnown],encodings)
                    match = face_recognition.compare_faces([student.encoding],face_encoding,tolerance=0.6)
                    if match:
                        present_students.append(student)
            #for student in present_students:
                row = [student.name,today.strftime("%Y-%m-%d"),"Present"]
                ws.append(row)
                wb.save("attendance.xlsx")
                return(f"{student.name}is present")
       
                    

            #return(f"{matches.name} is present")
        cv2.imshow('Video', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        #return ("Marked Succesfully !")
        #return(f"{matches[0].name} is present")                
    return redirect (url_for('index'))

@app.route("/download_attendance")
def download_attendance():
    # Send the Excel file to the user as a download
    return send_file("attendance.xlsx", as_attachment=True)
if __name__ == '__main__':
    db.create_all()
    app.run(debug=True)